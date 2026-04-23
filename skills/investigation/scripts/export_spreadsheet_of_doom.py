#!/usr/bin/env python3
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=14)
BODY_FONT = Font(size=14)
ALT_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
SHEET_ORDER = [
    ("timeline.csv", "Timeline"),
    ("systems.csv", "Systems"),
    ("users.csv", "Users"),
    ("host-indicators.csv", "Host Indicators"),
    ("network-indicators.csv", "Network Indicators"),
    ("task-tracker.csv", "Task Tracker"),
    ("evidence-tracker.csv", "Evidence Tracker"),
    ("keywords.csv", "Keywords"),
]


def autofit(width_values: list[list[str]]) -> list[int]:
    widths: list[int] = []
    if not width_values:
        return widths
    col_count = max(len(row) for row in width_values)
    for idx in range(col_count):
        max_len = 0
        for row in width_values:
            value = row[idx] if idx < len(row) else ""
            if value is None:
                value = ""
            value_len = max(len(part) for part in str(value).splitlines() or [""])
            max_len = max(max_len, value_len)
        widths.append(min(max(max_len + 2, 12), 48))
    return widths


def load_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.reader(fh))


def build_sheet(workbook: Workbook, sheet_name: str, rows: list[list[str]]) -> None:
    ws = workbook.create_sheet(title=sheet_name[:31])
    for row in rows:
        ws.append(row)

    if rows:
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP_ALIGNMENT

        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = ALT_FILL
            for cell in ws[row_idx]:
                cell.font = BODY_FONT
                cell.alignment = WRAP_ALIGNMENT

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for idx, width in enumerate(autofit(rows), start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: export_spreadsheet_of_doom.py <sod_dir> <output_xlsx>", file=sys.stderr)
        return 1

    sod_dir = Path(sys.argv[1]).resolve()
    output_xlsx = Path(sys.argv[2]).resolve()

    if not sod_dir.is_dir():
        print(f"Spreadsheet of Doom directory not found: {sod_dir}", file=sys.stderr)
        return 1

    workbook = Workbook()
    workbook.remove(workbook.active)

    handled = set()
    for filename, sheet_name in SHEET_ORDER:
        csv_path = sod_dir / filename
        if csv_path.exists():
            rows = load_csv(csv_path)
            build_sheet(workbook, sheet_name, rows)
            handled.add(csv_path.name)

    for csv_path in sorted(sod_dir.glob("*.csv")):
        if csv_path.name in handled:
            continue
        rows = load_csv(csv_path)
        build_sheet(workbook, csv_path.stem.replace("-", " ").title(), rows)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
